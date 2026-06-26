import io
import json
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from app.repositories.patent_repository import PatentRepository
from app.models.patent import Patent
from app.core.logging import logger
from typing import Optional


class ReportingService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.patent_repo = PatentRepository(db)

    async def generate_pdf_report(self, patent_ids: list[str],
                                   include_ai_summary: bool = True,
                                   include_claims: bool = True,
                                   include_metadata: bool = True) -> bytes:
        patents = []
        for pid in patent_ids:
            p = await self.patent_repo.get(pid)
            if p:
                patents.append(p)

        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, ListFlowable, ListItem,
        )
        from reportlab.lib import colors

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, spaceAfter=12,
        )
        heading_style = ParagraphStyle(
            'CustomHeading', parent=styles['Heading2'],
            fontSize=14, spaceAfter=6, spaceBefore=12,
        )
        normal_style = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=10, leading=14,
        )

        elements.append(Paragraph("Patent Intelligence Report", title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            normal_style
        ))
        elements.append(Spacer(1, 0.25 * inch))

        for i, patent in enumerate(patents):
            if i > 0:
                elements.append(PageBreak())

            elements.append(Paragraph(f"Patent {i + 1}: {patent.title or 'Untitled'}", heading_style))
            elements.append(Paragraph(
                f"Patent Number: {patent.patent_number or 'N/A'}", normal_style
            ))

            if include_metadata:
                elements.append(Spacer(1, 0.1 * inch))
                elements.append(Paragraph("Metadata", heading_style))

                metadata = [
                    ["Field", "Value"],
                    ["Patent Number", patent.patent_number or "N/A"],
                    ["Assignee", patent.assignee or "N/A"],
                    ["Inventors", ", ".join(patent.inventors or []) or "N/A"],
                    ["Filing Date", str(patent.filing_date or "N/A")],
                    ["Publication Date", str(patent.publication_date or "N/A")],
                    ["Status", patent.status or "N/A"],
                ]

                if patent.cpc_classifications:
                    metadata.append([
                        "CPC Classifications",
                        ", ".join(patent.cpc_classifications)
                    ])
                if patent.ipc_classifications:
                    metadata.append([
                        "IPC Classifications",
                        ", ".join(patent.ipc_classifications)
                    ])

                t = Table(metadata, colWidths=[1.5 * inch, 4.5 * inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('PADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)

            if patent.abstract:
                elements.append(Spacer(1, 0.1 * inch))
                elements.append(Paragraph("Abstract", heading_style))
                elements.append(Paragraph(patent.abstract, normal_style))

            if include_ai_summary and patent.abstract:
                elements.append(Spacer(1, 0.1 * inch))
                elements.append(Paragraph("AI Analysis", heading_style))
                elements.append(Paragraph(
                    f"Technology domain and key innovations extracted from patent content.",
                    normal_style
                ))

            if include_claims and patent.claims:
                elements.append(Spacer(1, 0.1 * inch))
                elements.append(Paragraph("Claims", heading_style))
                claims_list = []
                for idx, claim in enumerate(
                    patent.claims if isinstance(patent.claims, list) else [str(patent.claims)]
                ):
                    claims_list.append(Paragraph(
                        f"<b>{idx + 1}.</b> {claim}",
                        normal_style
                    ))
                for item in claims_list[:20]:
                    elements.append(item)

            elements.append(Spacer(1, 0.15 * inch))
            elements.append(Paragraph(
                f"Citations: {len(patent.citations or [])} references",
                normal_style
            ))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    async def generate_xlsx_report(self, patent_ids: list[str]) -> bytes:
        patents = []
        for pid in patent_ids:
            p = await self.patent_repo.get(pid)
            if p:
                patents.append(p)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Patent Report"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = [
            "Patent Number", "Title", "Assignee", "Inventors",
            "Filing Date", "Publication Date", "Status",
            "CPC Classifications", "IPC Classifications",
            "Abstract", "Citation Count", "Claim Count",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, patent in enumerate(patents, 2):
            values = [
                patent.patent_number or "",
                patent.title or "",
                patent.assignee or "",
                ", ".join(patent.inventors or []),
                str(patent.filing_date or ""),
                str(patent.publication_date or ""),
                patent.status or "",
                ", ".join(patent.cpc_classifications or []),
                ", ".join(patent.ipc_classifications or []),
                patent.abstract or "",
                len(patent.citations or []),
                len(patent.claims) if isinstance(patent.claims, list) else 0,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 60
        ws.column_dimensions['K'].width = 14
        ws.column_dimensions['L'].width = 14

        ws2 = wb.create_sheet("Summary")
        ws2.cell(row=1, column=1, value="Metric").font = header_font
        ws2.cell(row=1, column=2, value="Value").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=2, column=1, value="Total Patents")
        ws2.cell(row=2, column=2, value=len(patents))
        ws2.cell(row=3, column=1, value="Report Generated")
        ws2.cell(row=3, column=2, value=datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC'))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
